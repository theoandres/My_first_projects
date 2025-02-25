from openpyxl import Workbook, load_workbook
from pathlib import Path


class Task:
    def __init__(self, name, description, priority):
        self.name = name
        self.description = description
        self.priority = priority
        self.completed = False


    def complete_task(self):
        self.completed = True
        print('Task was completed successfully')


class TaskManager:
    def __init__(self):
        self.task_memory = []


    def add_task(self):
        name = input("Enter a name of the task: ")
        description = input("Enter a description of the task: ")
        while True:
             try:
                 priority = int(input("Enter a priority of the task (from 1 to 5]: "))
                 if 1 <= priority <= 5:
                     new_task = Task(name, description, priority)
                     self.task_memory.append(new_task)
                     print(f"Task '{name}' was added.")
                     break
             except ValueError:
                    print('Please enter valid priority and try again.')


    def complete_task(self):
        while True:
            if len(self.task_memory) > 0:
                self.show_tasks()
                index = input('Enter a number of task to complete: ')
                if 0 < int(index) <= len(self.task_memory):
                    self.task_memory[int(index)-1].complete_task()
                    break
                else:
                    print("Invalid input, please try again")
            else:
                print('No tasks available')
                break


    def show_tasks(self):
        if len(self.task_memory) >= 1:
            for idx, task in enumerate(self.task_memory, start=1):
                status = "completed" if task.completed else "not completed"
                print(f"{idx} : {task.name} - {task.description}, Priority: {task.priority}, Status: {status}")
        else:
            print('U have no tasks to do.')


    def save_and_exit(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "todolist_save"
        ws.append(['Name', 'Description', 'Priority', 'Completed'])
        for task in self.task_memory:
            ws.append([task.name, task.description, task.priority, task.completed])
        wb.save('todolistsave.xlsx')
        print('Tasks successfully saved')


    def remove_task(self):
        self.show_tasks()
        if len(self.task_memory) > 0:
            while True:
                try:
                    choice = int(input('Enter a number of task to delete: '))
                    if 1 <= choice <= len(self.task_memory):
                        self.task_memory.pop(choice - 1)
                        print('Task was successfully deleted.')
                        break
                    else:
                        print('Invalid input, please try again.')
                except ValueError:
                    print('Invalid input, please enter a valid value')
        else:
             print('No tasks to delete')


    def main_menu(self):
        while True:
            print("""
Choose of the following:
1 - Add task
2 - Show tasks
3 - Mark task as completed
4 - Remove task
5 - Save and exit
 """)
            choice = input()
            if choice == '1':
                self.add_task()
            elif choice == '2':
                self.show_tasks()
            elif choice == '3':
                self.complete_task()
            elif choice == '4':
                self.remove_task()
            elif choice == '5':
                self.save_and_exit()
                break
            else:
                print('Please enter a valid input')


if __name__ == "__main__":
    task_manager = TaskManager()
    file = Path('todolistsave.xlsx')
    if file.exists():
        wb = load_workbook(filename='todolistsave.xlsx')
        ws = wb.active
        for row in ws.iter_rows(min_row = 2, values_only = True):
            task = Task(name = row[0], description = row[1], priority = row[2])
            task_manager.task_memory.append(task)
    else:
        task_manager.task_memory = []
        print('Task manager is empty')

    task_manager.main_menu()




