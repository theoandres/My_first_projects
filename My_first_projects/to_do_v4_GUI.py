import tkinter
from  tkinter  import  *
from tkinter import messagebox
from tkinter.ttk import Treeview

from openpyxl import Workbook, load_workbook
from pathlib import Path

class Task:
    def __init__(self, name, description, priority):
        self.name = name
        self.description = description
        self.priority = priority
        self.completed = False


    def complete_task(self):
        self.completed = True


class TaskManager:
    def __init__(self, root):
        self.task_memory = []
        self.root = root
        self.root.iconbitmap("todo_icon.ico")
        self.root.title('To-Do App by Theo')


    def add_task(self):
        add_task_window = Toplevel(self.root)
        add_task_window.geometry('400x350')
        add_task_window.iconbitmap('todo_icon.ico')

        # Name part

        Label(add_task_window, text='Task Name:').pack(anchor="w", padx=10)
        name_entry = Entry(add_task_window, width=40, font=("Arial", 10))
        name_entry.pack(padx=10, pady=5, fill="x")


        # Description part
        Label(add_task_window, text='Description:').pack(anchor="w", padx=10)
        desc_text = Text(add_task_window, height=4, width=40, font=("Arial", 10), wrap="word")
        desc_text.pack(padx=10, pady=5, fill="x")


        # Priority part
        Label(add_task_window, text='Priority (1-5):').pack(anchor="w", padx=10)
        priority_entry = Entry(add_task_window, width=5, font=("Arial", 10))
        priority_entry.pack(padx=10, pady=5, anchor = 'w')


        # Funkcion for saving tasks
        def save_new_task():
            name = name_entry.get()
            description = desc_text.get("1.0", "end-1c")
            priority_text = priority_entry.get()
            if len(name) > 0:
                try:
                     priority = int(priority_text)
                     if priority < 1 or priority > 5:
                         messagebox.showerror("Valid priority", "ERROR: Priority must be a number between 1 and 5.")
                         priority_entry.delete(0, "end")
                         return
                     else:
                         new_task = Task(name, description, priority)
                         self.task_memory.append(new_task)
                         messagebox.showinfo("Task added", "Task was successfully added!")
                         add_task_window.destroy()  # Closes the window
                except ValueError:
                        messagebox.showerror("Valid priority", "ERROR: Priority must be a number.")
                        priority_entry.delete(0, "end")
                        return
            else:
                messagebox.showerror("Name missing", "Please enter a name of task.")
                return

        # Saving button
        save_task = Button(add_task_window, text="Save task", width=30, command= lambda: save_new_task(self))
        save_task.pack(pady=20)

        back_button = Button(add_task_window, text='Back to main menu', width=30,
                             command=lambda: back_main_from_show_tasks(self))
        back_button.pack()

        def back_main_from_show_tasks(self):
            add_task_window.destroy()

    # Function for GUI task interface. used to save space
    def tasks_GUI(self):
        self.task_window = Toplevel()
        self.task_treeview = Treeview(self.task_window, columns=('Name', 'Description', 'Status', 'Priority'),
                                      show='headings')
        self.task_window.iconbitmap("todo_icon.ico")
        self.task_treeview.heading("Priority", text="Priority")
        self.task_treeview.heading("Name", text="Task Name")
        self.task_treeview.heading("Description", text="Description")
        self.task_treeview.heading("Status", text="Status")
        verscrlbar = Scrollbar(self.task_window,
                               orient="vertical",
                               command=self.task_treeview.yview)
        verscrlbar.pack(side='right', fill='y')
        self.task_treeview.pack()
        self.task_treeview.configure(xscrollcommand=verscrlbar.set)

        for column in self.task_treeview["columns"]:
            self.task_treeview.column(column, anchor=CENTER)  # This will center text in rows
            self.task_treeview.heading(column, text=column)

        if len(self.task_memory) >= 1:
            for idx, task in enumerate(self.task_memory, start=1):
                Status = "completed" if task.completed else "not completed"
                self.task_treeview.insert('', END, values=(task.name, task.description, Status, task.priority))
        else:
            self.task_treeview.insert("", "end", values=("", "No tasks available", "", ""))


    def complete_task(self):
        self.tasks_GUI()
        back_button = Button(self.task_window, text='Back to main menu',
                             command=lambda: back_main(self))
        complete_button = Button(self.task_window, text='Complete task',
                                 command=lambda: complete_item(self))
        complete_button.pack()


        def complete_item(self):
            selected_item = self.task_treeview.selection()
            iid = selected_item[0]
            task_values = self.task_treeview.item(iid, "values")
            task_name = task_values[0]
            for task in self.task_memory:
                if task.name == task_name:
                    task.completed = True
            self.task_treeview.delete(iid)
            self.task_treeview.insert("", "end", values=(task_name, task_values[1], "completed", task_values[3]))
            messagebox.showinfo('Task completed', "Task was successfully completed.")


        def back_main (self):
            self.task_window.destroy()
        back_button.pack(pady=20)


    def show_tasks(self):
        show_task_window = Toplevel()
        show_task_window.geometry('900x300')
        show_task_window.iconbitmap("todo_icon.ico")
        task_treeview = Treeview(show_task_window, columns=('Name','Description','Status', 'Priority'), show='headings')
        task_treeview.heading("Priority", text = "Priority")
        task_treeview.heading("Name", text = "Task Name")
        task_treeview.heading("Description", text = "Description")
        task_treeview.heading("Status", text = "Status")
        verscrlbar = Scrollbar(show_task_window,
                                   orient="vertical",
                                   command=task_treeview.yview)
        verscrlbar.pack(side = 'right', fill = 'y')
        task_treeview.pack()
        task_treeview.configure(xscrollcommand=verscrlbar.set)

        for column in task_treeview["columns"]:
            task_treeview.column(column, anchor=CENTER)  # This will center text in rows
            task_treeview.heading(column, text=column)

        if len(self.task_memory) >= 1:
            for idx, task in enumerate(self.task_memory, start=1):
                status = "completed" if task.completed else "not completed"
                task_treeview.insert('', END, values = (task.name, task.description, status, task.priority))
        else:
            task_treeview.insert("", "end", values=("", "No tasks available", "", ""))
        back_button = Button(show_task_window, text='Back to main menu', command = lambda: back_main_from_show_tasks(self))
        back_button.pack(pady = 20)

        def back_main_from_show_tasks (self):
            show_task_window.destroy()


    def save_and_exit(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "todolist_save"
        ws.append(['Name', 'Description', 'Priority', 'Completed'])
        for task in self.task_memory:
            ws.append([task.name, task.description, task.priority, task.completed])
        wb.save('todolistsave.xlsx ')
        messagebox.showinfo('Save', "To-Do App was successfully saved. See u soon!")
        exit()

    def remove_task(self):
        self.tasks_GUI()
        back_button = Button(self.task_window, text='Back to main menu',
                             command=lambda: back_main(self))
        delete_button = Button(self.task_window, text='Delete task',
                               command=lambda: remove_item(self))


        def remove_item ():
            selected_item = self.task_treeview.selection()
            iid = selected_item[0]
            task_values = self.task_treeview.item(iid, "values")
            task_name = task_values[0]
            self.task_memory = [task for task in self.task_memory if task.name != task_name]
            self.task_treeview.delete(selected_item)
            messagebox.showinfo('Task deleted', "Task was successfully deleted.")

        delete_button.pack()


        def back_main (self):
            self.task_window.destroy()
        back_button.pack(pady=20)


    def main_menu(self):
        menu_window = Toplevel(self.root)
        menu_window.geometry('500x300')
        menu_window.iconbitmap("todo_icon.ico")
        main_menu_text = Label(menu_window, text = 'Choose of the following options:')
        add_task = Button(menu_window, text="Add task", width=30, command=self.add_task)
        show_task = Button(menu_window, text="Show tasks", width=30, command=self.show_tasks)
        complete_task = Button(menu_window, text="Mark task as completed", width=30, command=self.complete_task)
        remove_task = Button(menu_window, text="Remove task", width=30, command=self.remove_task)
        save_exit = Button(menu_window, text="Save and exit", width=30, command=self.save_and_exit)
        main_menu_text.pack(pady = 20)
        add_task.pack(pady = 5)
        show_task.pack(pady = 5)
        complete_task.pack(pady = 5)
        remove_task.pack(pady = 5)
        save_exit.pack(pady = 5)

        tkinter.mainloop()

if __name__ == "__main__":
    root = Tk()  # Vytvoření hlavního okna mimo třídu
    root.withdraw()  # Skryje hlavní okno (root) – nebude viditelné, jen existuje
    task_manager = TaskManager(root)
    file = Path('todolistsave.xlsx')
    if file.exists():
        wb = load_workbook(filename='todolistsave.xlsx')
        ws = wb.active
        for row in ws.iter_rows(min_row = 2, values_only = True):
            task = Task(name = row[0], description = row[1], priority = row[2])
            task_manager.task_memory.append(task)
    else:
        task_manager.task_memory = []
    task_manager.main_menu()




